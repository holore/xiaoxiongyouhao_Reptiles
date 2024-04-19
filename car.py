import random
import re
import json
import time
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 构造请求头
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko)'
                         ' Chrome/100.0.4896.127 Safari/537.36'}

# 代理ip
proxy = "127.0.0.1:7890"
proxies = {'http': "http://" + proxy,
           'https': 'http://' + proxy
           }

# 编译正则表达式
satisfaction_pattern = re.compile(r"text: '(.+?)车主满意度'")
score_pattern = re.compile(r"subtitle: {\s*text: '(\d+)个车主综合评分: (\d+(?:\.\d+)?)'")
various_score_pattern = re.compile(r'"data":(\[.*?])')

recommendation_pattern = re.compile(r""""y":(\d+).*?\\u63a8\\u8350.*?(\d+)""")
brand_repurchase_intention_pattern = re.compile(r""""y":(\d+).*?\\u4f1a\\u518d\\u6b21\\u8d2d\\u4e70.*?(\d+)""")


# 解析网页内容
def parse_page(url):
    rsp = requests.get(url, proxies=proxies, headers=headers)
    if rsp.status_code == 200:
        return rsp.text
    else:
        print(f"Failed to fetch data from {url}")
        return None


# 计算百分比
def calculate_percentage(data):
    count, score = map(int, data)
    total = count + score
    return round(count * 100 / total, 2), total


# 提取数据并填入表格
def extract_and_fill(ws, cell_coordinate, text):
    if "阳春白雪" in text:
        print(f"No data available for {cell_coordinate}")
        ws[cell_coordinate.replace('B', 'C')] = "No data available"
        return

    satisfaction_match = satisfaction_pattern.findall(text)
    score_match = score_pattern.findall(text)
    various_score_match = various_score_pattern.search(text)
    recommendation_data = recommendation_pattern.findall(text)
    brand_repurchase_intention_data = brand_repurchase_intention_pattern.findall(text)

    if satisfaction_match:
        print("车主满意度:", satisfaction_match)
    if score_match:
        print("车型评分:", score_match)
        ws[cell_coordinate.replace('B', 'C')] = float(score_match[0][0])
        ws[cell_coordinate.replace('B', 'D')] = float(score_match[0][1])
    if various_score_match:
        various_score = json.loads(various_score_match.group(1))
        print("车型各项评分:", various_score)
        for i, score in enumerate(various_score):
            ws[cell_coordinate.replace('B', get_column_letter(i + 5))] = float(score)

    # 计算推荐度总评分人数和百分比
    if recommendation_data:
        print("推荐度：", recommendation_data[0])
        ws[cell_coordinate.replace('B', 'O')], ws[cell_coordinate.replace('B', 'P')] = calculate_percentage(
            recommendation_data[0])

    # 计算品牌复购意向总评分人数和百分比
    if brand_repurchase_intention_data:
        print('品牌复购:', brand_repurchase_intention_data)
        ws[cell_coordinate.replace('B', 'Q')], ws[cell_coordinate.replace('B', 'R')] = calculate_percentage(
            brand_repurchase_intention_data[0])

    time.sleep(random.randint(3, 7))


# 读取小熊油耗汽车链接
wb = load_workbook('车系信息1.xlsx')
ws = wb['小熊油耗信息']

try:
    count = 0
    for cell in ws['B']:
        # if cell.value == 'https://www.xiaoxiongyouhao.com/chexiyh/2841.html':
        #     break
        if cell.value:
            url = cell.value
            cell_coordinate = cell.coordinate

            # 如果C列已经存在数据，则跳过该单元格
            if ws[cell_coordinate.replace('B', 'C')].value is not None:
                print(f"Skipping {cell_coordinate}: Data already exists.")
                continue

            page_text = parse_page(url)
            if page_text:
                extract_and_fill(ws, cell_coordinate, page_text)

        count += 1
        if count % 100 == 0:
            print('保存')
            wb.save("车系信息1.xlsx")
            break


finally:
    wb.save("车系信息1.xlsx")
